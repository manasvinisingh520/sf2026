"""
Utility functions for reading and processing single-cell data.

This module provides helper functions for:
- Reading .mtx files (Matrix Market format)
- Creating AnnData objects from matrices
- Reading Excel files into pandas DataFrames
- Filtering cells based on metadata criteria
- Getting file paths for region-specific data
"""

import scipy.io
from scipy.sparse import csr_matrix
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, List, Dict, Union, Tuple
import anndata
from collections import defaultdict
import re
import hashlib
import pickle
import os
import config as cfg


def read_mtx_file(mtx_path, row_annotation_path=None, col_annotation_path=None, 
                  transpose=True, make_gene_names_unique=True):
    """
    Read a .mtx file and optionally load row/column annotations.
    
    Parameters:
    -----------
    mtx_path : str
        Path to the .mtx file
    row_annotation_path : str, optional
        Path to file with row names (genes/features). 
        Each line should contain one name.
    col_annotation_path : str, optional
        Path to file with column names (cell barcodes).
        Each line should contain one name.
    transpose : bool, default=True
        If True, transpose the matrix so rows=genes, cols=cells.
        MTX format typically stores cells as rows and genes as columns.
    make_gene_names_unique : bool, default=True
        If True, ensure gene names are unique by appending _1, _2, etc.
    
    Returns:
    --------
    matrix : scipy.sparse.csr_matrix or scipy.sparse.csc_matrix
        Sparse matrix representation
    gene_names : array-like, optional
        Gene/feature names if row_annotation_path provided
    cell_names : array-like, optional
        Cell barcode names if col_annotation_path provided
    """
    
    # Read the .mtx file using scipy
    # This returns a COO (Coordinate) format sparse matrix
    matrix = scipy.io.mmread(mtx_path)
    print(f"Loaded matrix shape: {matrix.shape}")
    print(f"Number of non-zero entries: {matrix.nnz:,}")
    print(f"Sparsity: {(1 - matrix.nnz / (matrix.shape[0] * matrix.shape[1])) * 100:.2f}%")
    
    # Transpose if needed (MTX format often has cells as rows, genes as columns)
    if transpose:
        matrix = matrix.T
        print(f"Transposed matrix shape: {matrix.shape}")
    
    # Convert to CSR format for efficient row operations (or CSC for column operations)
    # CSR is typically preferred for most operations
    matrix = csr_matrix(matrix)
    
    # Load row annotations (genes/features) if provided
    gene_names = None
    if row_annotation_path:
        # Read gene names from file
        # Handle different formats (quoted strings, CSV, TSV, etc.)
        try:
            with open(row_annotation_path, 'r') as f:
                gene_names = [line.strip().strip('"') for line in f]
            
            # Handle unique gene names (common issue in scRNA-seq)
            if make_gene_names_unique:
                unique_names = []
                name_counts = {}
                for name in gene_names:
                    if name in name_counts:
                        name_counts[name] += 1
                        unique_names.append(f"{name}_{name_counts[name]}")
                    else:
                        name_counts[name] = 0
                        unique_names.append(name)
                gene_names = unique_names
            
            print(f"Loaded {len(gene_names)} gene names")
            
        except Exception as e:
            print(f"Warning: Could not load row annotations: {e}")
    
    # Load column annotations (cell barcodes) if provided
    cell_names = None
    if col_annotation_path:
        try:
            with open(col_annotation_path, 'r') as f:
                cell_names = [line.strip().strip('"') for line in f]
            print(f"Loaded {len(cell_names)} cell barcodes")
        except Exception as e:
            print(f"Warning: Could not load column annotations: {e}")
    
    return matrix, gene_names, cell_names


def read_excel_columns(
    file_path: str,
    columns: Optional[List[str]] = None,
    sheet_name: Optional[Union[str, int]] = 0,
    dtype: Optional[Dict[str, str]] = None,
    engine: Optional[str] = None,
    use_progress: bool = True
) -> pd.DataFrame:
    """
    Read a Microsoft Excel file and return a pandas DataFrame with selected columns.
    Results are cached based on all function arguments to speed up repeated calls.

    Parameters:
    -----------
    file_path : str
        Path to the Excel file (.xlsx, .xls, .xlsm).
    columns : list of str, optional
        Column names to select. If None, returns all columns.
    sheet_name : str or int, optional (default=0)
        Name or index of the sheet to read.
    dtype : dict, optional
        Optional mapping of column name -> dtype to enforce after reading.
    engine : str, optional
        Excel engine to use (e.g., 'openpyxl'). If None, pandas will auto-select.
    use_progress : bool, optional (default=True)
        If True, show a progress bar while reading using openpyxl+tqdm when available.

    Returns:
    --------
    df : pandas.DataFrame
        DataFrame containing the selected columns.
    """
    # Create cache directory if it doesn't exist
    cache_dir = Path(".cache")
    cache_dir.mkdir(exist_ok=True)
    
    # Generate cache key from all function arguments
    cache_key_parts = [
        str(file_path),
        str(sorted(columns) if columns else None),
        str(sheet_name),
        str(sorted(dtype.items()) if dtype else None),
        str(engine),
        str(use_progress)
    ]
    cache_key_str = "|".join(cache_key_parts)
    cache_key_hash = hashlib.md5(cache_key_str.encode()).hexdigest()
    cache_file = cache_dir / f"excel_cache_{cache_key_hash}.pkl"
    
    # Check if source file exists and get its modification time
    source_mtime = None
    if os.path.exists(file_path):
        source_mtime = os.path.getmtime(file_path)
    
    # Try to load from cache
    if cache_file.exists():
        try:
            with open(cache_file, 'rb') as f:
                cached_data = pickle.load(f)
                cached_mtime = cached_data.get('source_mtime')
                cached_df = cached_data.get('df')
                
                # Check if source file hasn't been modified since cache was created
                if cached_mtime is not None and source_mtime is not None:
                    if cached_mtime >= source_mtime:
                        return cached_df.copy()
                elif cached_mtime == source_mtime:  # Both None or same
                    return cached_df.copy()
        except Exception as e:
            # If cache loading fails, continue to read from file
            print(f"Warning: Failed to load cache ({e}), reading from file...")
    
    # Cache miss or invalid - read from file
    # If requested, try progress-enabled path using openpyxl + tqdm
    if use_progress:
        try:
            import openpyxl  # type: ignore
            try:
                from tqdm import tqdm  # type: ignore
            except Exception:
                tqdm = None  # fallback without visible progress

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = list(header_row)

            # Determine which columns to keep
            if columns is not None:
                name_to_idx = {name: idx for idx, name in enumerate(headers)}
                missing = [c for c in columns if c not in name_to_idx]
                if missing:
                    print(f"Warning: Missing columns not found in Excel: {missing}")
                selected = [c for c in columns if c in name_to_idx]
                selected_indices = [name_to_idx[c] for c in selected]
                out_headers = selected
            else:
                selected_indices = list(range(len(headers)))
                out_headers = headers

            total_rows = max(ws.max_row - 1, 0)
            progress_iter = tqdm(total=total_rows, desc="Reading Excel", unit="row") if tqdm else None

            records = []
            # Iterate data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    if progress_iter:
                        progress_iter.update(1)
                    continue
                # Ensure row has enough length
                row_values = list(row)
                if len(row_values) < len(headers):
                    row_values += [None] * (len(headers) - len(row_values))
                # Select indices
                record = [row_values[i] for i in selected_indices]
                records.append(record)
                if progress_iter:
                    progress_iter.update(1)

            if progress_iter:
                progress_iter.close()
            wb.close()

            df = pd.DataFrame.from_records(records, columns=out_headers)
        except Exception as e:
            print(f"Note: Falling back to pandas.read_excel (reason: {e})")
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
    else:
        # Read the sheet via pandas directly
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)

    # Select requested columns if provided
    if columns is not None:
        missing = [c for c in columns if c not in df.columns]
        if missing:
            print(f"Warning: Missing columns not found in Excel: {missing}")
        present = [c for c in columns if c in df.columns]
        df = df[present]

    # Apply dtype conversions if requested
    if dtype:
        for col, target_dtype in dtype.items():
            if col in df.columns:
                try:
                    df[col] = df[col].astype(target_dtype)
                except Exception as e:
                    print(f"Warning: Could not cast column '{col}' to {target_dtype}: {e}")

    # Save to cache
    try:
        with open(cache_file, 'wb') as f:
            pickle.dump({
                'df': df,
                'source_mtime': source_mtime
            }, f)
    except Exception as e:
        # If caching fails, continue without cache
        print(f"Warning: Failed to save cache ({e})")

    return df


def create_anndata_object(matrix, gene_names=None, cell_names=None, obs=None, transpose=False):
    """
    Create an AnnData object (standard format for single-cell analysis).
    Requires scanpy/anndata package.
    
    Parameters:
    -----------
    matrix : scipy.sparse matrix
        Expression matrix (assumed to be genes × cells if transpose=False)
    gene_names : array-like, optional
        Gene names
    cell_names : array-like, optional
        Cell barcodes
    obs : pandas.DataFrame, optional
        Cell metadata (must have same number of rows as cells)
    transpose : bool, default=False
        If True, transpose the matrix before creating AnnData
        (AnnData expects cells as rows, genes as columns)
    
    Returns:
    --------
    adata : anndata.AnnData
        AnnData object ready for analysis
    """
    try:
        import anndata as ad
        
        # Create AnnData object
        # AnnData expects cells as rows (obs), genes as columns (var)
        if transpose:
            adata = ad.AnnData(matrix.T)
        else:
            # Matrix is already in cells × genes format
            adata = ad.AnnData(matrix)
        
        if gene_names is not None:
            adata.var_names = gene_names
        if cell_names is not None:
            adata.obs_names = cell_names
        if obs is not None:
            adata.obs = obs
        
        return adata
    
    except ImportError:
        print("anndata package not installed. Install with: pip install anndata scanpy")
        return None

def filter_anndata_object(adata, min_genes=200, min_cells=3, min_counts=None, 
                          max_counts=None, mito_max=None, RIN_threshold=None):
    """
    Filter an AnnData object based on the number of genes and cells.
    
    Parameters:
    -----------
    adata : anndata.AnnData
        AnnData object to filter
    min_genes : int, default=200
        Minimum number of genes detected per cell
    min_cells : int, default=3
        Minimum number of cells expressing each gene (0 = no gene filtering)
    min_counts : int, optional
        Minimum total counts (UMIs) per cell
    max_counts : int, optional
        Maximum total counts (UMIs) per cell    
    Returns:
    --------
    adata : anndata.AnnData
        Filtered AnnData object
    """
    import scanpy as sc
    sc.pp.filter_cells(adata, min_genes=min_genes, min_counts=min_counts, max_counts=max_counts)
    # Only filter genes if min_cells > 0
    if min_cells > 0:
        sc.pp.filter_genes(adata, min_cells=min_cells)
    if mito_max is not None:
        mask = (adata.obs['percent.mito'] < mito_max).fillna(False)
        adata = adata[mask].copy()
    if RIN_threshold is not None:
        # Exclude cells where RIN == 'unk', then convert to numeric for comparison
        mask_unk = (adata.obs['RIN'] != 'unk')
        rin_numeric = pd.to_numeric(adata.obs['RIN'], errors='coerce')
        mask = mask_unk & (rin_numeric > RIN_threshold).fillna(False)
        adata = adata[mask].copy()
    else:
        print("Warning: mito_max is not set. Skipping mito filter.")

    return adata


# Shuffle an AnnData object based on the seed
def shuffle_data(adata: anndata.AnnData, seed: int) -> anndata.AnnData:
    rng = np.random.RandomState(seed)
    n_cells = adata.n_obs
    shuffled_indices = rng.permutation(n_cells)
    # Create a new AnnData object with shuffled data by reordering underlying arrays
    # This avoids issues with integer vs string observation names
    shuffled_adata = anndata.AnnData(
        X=adata.X[shuffled_indices] if adata.X is not None else None,
        obs=adata.obs.iloc[shuffled_indices].copy(),
        var=adata.var.copy(),
        obsm={k: v[shuffled_indices] for k, v in adata.obsm.items()} if adata.obsm else {},
        varm=adata.varm.copy() if adata.varm else {},
        uns=adata.uns.copy() if adata.uns else {},
        layers={k: v[shuffled_indices] for k, v in adata.layers.items()} if adata.layers else {}
    )
    return shuffled_adata


def aggregate_cells_into_pseudobulk(adata, target_cells_per_bin, filter_patients_cell_threshold=80, seed=100, patient_col='SampleName'):
    """
    Aggregate cells into pseudobulk samples.
    
    CRITICAL: Aggregate cells into pseudobulk samples if dataset is too large
    DESeq2 is designed for bulk RNA-seq with few samples (3-10 per condition)
    Single-cell data with thousands of cells causes memory errors
    Pseudobulk aggregation is standard practice for single-cell DGE analysis
    
    Parameters:
    -----------
    adata : AnnData
        AnnData object with cells to aggregate
    target_cells_per_bin : int
        Target number of cells per bin
    filter_patients_cell_threshold : int, default=80
        Minimum number of cells per patient to include
    seed : int, default=100
        Random seed for shuffling cells
    patient_col : str, default='SampleName'
        Column name in adata.obs containing patient/donor identifiers
        
    Returns:
    --------
    adata_pseudobulk : AnnData
        AnnData object with pseudobulk samples
    """
    from scipy.sparse import csr_matrix
    from anndata import AnnData
        
    pseudobulk_data = []
    pseudobulk_metadata = []
    
    if patient_col not in adata.obs.columns:
        raise ValueError(f"Patient column '{patient_col}' not found in adata.obs. Available columns: {list(adata.obs.columns)}")
    
    unique_patients = adata.obs[patient_col].dropna().unique()
    total_bins = 0
                
    for patient_name in unique_patients:
        patient_mask = adata.obs[patient_col] == patient_name
        patient_cells = adata[patient_mask]
        num_cells = patient_cells.n_obs
            
        if num_cells < filter_patients_cell_threshold:
            print(f"  {patient_name}: {num_cells} cells (skipped, < {filter_patients_cell_threshold} cells)")
            continue
            
        # Shuffle cells for this patient using the seed
        if seed != 100:
            patient_cells = shuffle_data(patient_cells, seed)
            
        num_bins = max(1, int(np.ceil(num_cells / target_cells_per_bin)))
        total_bins += num_bins
        print(f"  {patient_name}: {num_cells} cells {num_bins} bins")
        
        # Get patient metadata (same across all cells for this patient)
        # Use first cell's values since they're consistent across all cells
        patient_metadata = patient_cells.obs.iloc[0].to_dict()
        
        # Distribute cells evenly across bins
        base_cells_per_bin = num_cells // num_bins
        remainder = num_cells % num_bins
        
        cell_idx = 0
        for bin_idx in range(num_bins):
            bin_size = base_cells_per_bin + (1 if bin_idx < remainder else 0)
            start_idx = cell_idx
            end_idx = cell_idx + bin_size
            cell_idx = end_idx
            
            # Work directly with underlying arrays to avoid AnnData indexing issues
            # Get the subset of the count matrix directly
            X_subset = patient_cells.X[start_idx:end_idx]
            
            # Calculate bin counts (sum across cells, keep gene dimension)
            if hasattr(X_subset, 'toarray'):
                # Sparse matrix
                bin_counts = X_subset.sum(axis=0).A1
            else:
                # Dense matrix
                bin_counts = X_subset.sum(axis=0)
            
            # Build metadata dictionary with all patient-level metadata
            metadata_dict = patient_metadata.copy()
            metadata_dict['_sample_id'] = f"{patient_name}_bin{bin_idx+1}"
            
            pseudobulk_data.append(bin_counts)
            pseudobulk_metadata.append(metadata_dict)
    
    print(f"  Total bins: {total_bins}")
    print(f"\nTotal pseudobulk samples: {len(pseudobulk_data)}")
    
    pseudobulk_matrix = np.array(pseudobulk_data)
    pseudobulk_obs = pd.DataFrame(pseudobulk_metadata)
    pseudobulk_obs.index = pseudobulk_obs['_sample_id']
    
    adata_pseudobulk = AnnData(
        X=csr_matrix(pseudobulk_matrix),
        obs=pseudobulk_obs,
        var=adata.var.copy()
    )
    
    return adata_pseudobulk



def get_region_file_paths(region, data_dir="data", base_prefix="2025-10-22_Astrocytes_{region}"):
    """
    Get file paths for region-specific matrix, row, and column annotation files.
    
    Parameters:
    -----------
    region : str
        Region name (e.g., 'EC', 'ITG', 'PFC', 'V1', 'V2')
    data_dir : str or Path, default="data"
        Base directory containing region files
    base_prefix : str, default="2025-10-22_Astrocytes_{region}"
        File naming pattern with {region} placeholder
    
    Returns:
    --------
    mtx_path : Path
        Path to matrix.mtx file
    row_path : Path
        Path to row_annotation.txt file
    col_path : Path
        Path to cell_annotation.txt file
    """
    data_dir = Path(data_dir)
    prefix = base_prefix.format(region=region)
    mtx_path = (data_dir / f"{prefix}_matrix.mtx").resolve()
    row_path = (data_dir / f"{prefix}_row_annotation.txt").resolve()
    col_path = (data_dir / f"{prefix}_cell_annotation.txt").resolve()
    return mtx_path, row_path, col_path


def get_dge_results_dir(version: int, base_dir: str = None) -> Path:
    """
    Get the path to DGE results directory for a given version.
    
    Parameters:
    -----------
    version : int
        DGE version (1, 2, 3, or 4)
    base_dir : str, optional
        Base directory for results. If None, uses current working directory.
    
    Returns:
    --------
    results_dir : Path
        Path to the DGE results directory (e.g., results/dge4)
    """
    if base_dir is None:
        from pathlib import Path
        base_dir = Path.cwd()
    else:
        base_dir = Path(base_dir)
    
    return base_dir / "results" / f"dge{version}"


def filter_cells(matrix, cell_names, metadata, mito_max=0.15, extra_filter = False):
    """
    Filter cells based on mitochondrial percentage threshold.
    
    Parameters:
    -----------
    matrix : scipy.sparse matrix
        Expression matrix (genes × cells)
    cell_names : list
        List of cell names/barcodes
    metadata : pandas.DataFrame
        Metadata DataFrame with 'cell_annotation' and 'percent.mito' columns
    mito_max : float, default=15.0
        Maximum mitochondrial percentage threshold
    
    Returns:
    --------
    filtered_matrix : scipy.sparse matrix
        Filtered expression matrix
    filtered_cell_names : list
        List of cell names that passed filters
    filtered_metadata : pandas.DataFrame
        Filtered metadata DataFrame
    """
    # Ensure cell_names matches matrix columns
    n_cols = matrix.shape[1]
    if len(cell_names) != n_cols:
        print(f"  Warning: cell_names length ({len(cell_names)}) != matrix columns ({n_cols})")
        cell_names = cell_names[:n_cols]  # Truncate to match matrix
    
    filtered_metadata = metadata[metadata['cell_annotation'].isin(cell_names)].copy()
    filtered_metadata = filtered_metadata.set_index('cell_annotation').reindex(cell_names)
    n_before_filter = len(filtered_metadata)

    # Debug: Check percent.mito values
    mito_values = filtered_metadata['percent.mito']
    print(f"  Mitochondrial percentage stats:")
    print(f"    Min: {mito_values.min():.2f}, Max: {mito_values.max():.2f}, Mean: {mito_values.mean():.2f}")
    print(f"    NaN count: {mito_values.isna().sum()}")
    print(f"    Threshold: {mito_max}")

    # Filter by mitochondrial percentage
    mito_mask = (filtered_metadata['percent.mito'] < mito_max).fillna(False)
    if extra_filter:
        # Convert to numeric, coercing errors to NaN
        RIN_numeric = pd.to_numeric(filtered_metadata['RIN'], errors='coerce')
        TotalGenes_numeric = pd.to_numeric(filtered_metadata['Total.Genes.Detected'], errors='coerce')
        RIN_genes_mask = ((RIN_numeric > 7.5) | (TotalGenes_numeric > 30000)).fillna(False)
        combined_mask = mito_mask & RIN_genes_mask
    else:
        combined_mask = mito_mask
    filtered_metadata = filtered_metadata[combined_mask].copy()
    
    print(f"  Before: {n_before_filter:,} cells")
    print(f"  After: {len(filtered_metadata):,} cells")
    print(f"  Filtered out: {n_before_filter - len(filtered_metadata):,} cells")
    
    # Get cell indices to keep (cells that passed filters)
    cells_to_keep = set(filtered_metadata.index.dropna().tolist())
    cell_indices_to_keep = [i for i, name in enumerate(cell_names) if name in cells_to_keep]
    
    # Ensure indices are within matrix bounds
    cell_indices_to_keep = [i for i in cell_indices_to_keep if 0 <= i < n_cols]
    
    # Filter matrix and cell names
    filtered_matrix = matrix[:, cell_indices_to_keep]
    filtered_cell_names = [cell_names[i] for i in cell_indices_to_keep]
    
    return filtered_matrix, filtered_cell_names, filtered_metadata


def min_max_normalize(matrix):
    mat = matrix.tocoo(copy=True).astype(float) #coordinate format
    # Convert to CSR for efficient row operations, then get max/min as dense arrays
    matrix_csr = matrix.tocsr()
    row_max = np.array(matrix_csr.max(axis=1).toarray()).ravel()
    row_min = np.array(matrix_csr.min(axis=1).toarray()).ravel()
    denom = row_max - row_min
    denom[denom == 0] = 1.0  # Avoid divide by zero
    mat.data = (mat.data - row_min[mat.row]) / denom[mat.row]
    return mat.tocsr()


def get_top_k_genes(df: pd.DataFrame, k: int, sort_by: str = 'padj') -> set:
    df_clean = df[df[sort_by].notna()].copy()
    if len(df_clean) == 0:
        return set()
    ascending = (sort_by == 'padj')
    df_sorted = df_clean.sort_values(sort_by, ascending=ascending)
    return list(df_sorted.head(k).index), set(df_sorted.head(k).index)

def get_DEGs(region, top_k=500, cell_type="Astrocytes", disable_intersection=False):
    """
    Get the intersection of top K genes (by padj) across all 6 comparisons for a region.
    
    Parameters:
    -----------
    region : str
        Region name (EC, ITG, PFC, V1, V2, CrossRegion)
    top_k : int
        Number of top genes to select from each file (sorted by padj)
    cell_type : str
        Cell type (Astrocytes or Microglia)
    disable_intersection : bool
        Whether to disable intersection of gene sets
    
    Returns:
    --------
    deg_genes : list
        List of gene names that are in the top K genes of all comparisons (intersection)
    """
    dge_results_dir="results/dge_final"
    dge_dir = Path(dge_results_dir)
    
    # Find all DGE results files for this region
    if cell_type == "Microglia":
        pattern = f"dge_results_microglia_{region}_*.csv"
    else:
        pattern = f"dge_results_{region}_*.csv"
    pattern = f"dge_results_{region}_*_1_vs_4.csv"
    dge_files = sorted(list(dge_dir.glob(pattern)))
    
    # Collect top K genes from each comparison
    gene_sets = []
    for dge_file in dge_files:
        try:
            df = pd.read_csv(dge_file, index_col=0)
            # Get top K genes sorted by padj
            top_genes_list, top_genes_set = get_top_k_genes(df, top_k, sort_by='padj')
            gene_sets.append(top_genes_set)
            print(f"  {dge_file.name}: {len(top_genes_set)} top genes")
            top_10_genes = top_genes_list[:min(len(top_genes_list), 10)]
            print(f"  First 10 genes: {top_10_genes}")
            print(f"  Gene types: {[cfg.gene_annotations_dict[gene] for gene in top_10_genes]}")

        except Exception as e:
            print(f"Warning: Could not load {dge_file}: {e}")
            continue
    
    if not gene_sets:
        print(f"No gene sets collected for region {region}")
        return []
 
    print(f"Total comparisons: {len(gene_sets)}")
    print(f"Gene set sizes: {[len(gs) for gs in gene_sets]}")
    
    # Return intersection of all top K gene sets
    # gene_sets is a list of sets, so we unpack with * to pass them as separate arguments
    if disable_intersection:
        common_genes = set.union(*gene_sets)
        print(f"Union size: {len(common_genes)}")

    else:
        common_genes = set.intersection(*gene_sets) if len(gene_sets) > 1 else gene_sets[0]
        print(f"Intersection size: {len(common_genes)}")

    ## Check intersection with genes from paper
    genes_from_paper = cfg.GENES_FROM_PAPER[region]
    common_genes_with_paper = common_genes.intersection(genes_from_paper)
    print(f"\033[92mIntersection with genes from paper: {len(common_genes_with_paper)} out of {len(genes_from_paper)}\033[0m")
    for gene in common_genes_with_paper:
        print(f"\033[92m{gene}\033[0m")
    
    return list(common_genes)
